from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from src.agents_flow.extraction_flow import OUTPUT_FIELDS


@dataclass(frozen=True)
class InputRecord:
    row_number: int
    nro_documento: str
    apellidos_nombres: str = ""

    @property
    def dni(self) -> str:
        return self.nro_documento


@dataclass(frozen=True)
class SearchResult:
    documento: str = ""
    tipo_documento: str = ""
    nombre: str = ""
    estado: str = ""
    nro_carne: str = ""
    modalidad: str = ""
    ruc: str = ""
    expediente: str = ""
    nro_expediente: str = ""
    anho_expediente: str = ""
    fecha_emision: str = ""
    fecha_vencimiento: str = ""
    empresa: str = ""
    curso_ruc_1: str = ""
    curso_razon_social_1: str = ""
    curso_evaluacion_1: str = ""
    curso_tipo_1: str = ""
    curso_fecha_inicio_1: str = ""
    curso_fecha_venc_1: str = ""
    curso_estado_1: str = ""
    curso_ruc_2: str = ""
    curso_razon_social_2: str = ""
    curso_evaluacion_2: str = ""
    curso_tipo_2: str = ""
    curso_fecha_inicio_2: str = ""
    curso_fecha_venc_2: str = ""
    curso_estado_2: str = ""
    licencia_numero: str = ""
    licencia_fecha_emision: str = ""
    licencia_fecha_venc: str = ""
    licencia_modalidad: str = ""
    licencia_restricciones: str = ""
    historial_ruc_1: str = ""
    historial_razon_social_1: str = ""
    historial_modalidad_1: str = ""
    historial_procedimiento_1: str = ""
    historial_fecha_emision_1: str = ""
    historial_fecha_venc_1: str = ""
    historial_fecha_baja_1: str = ""
    historial_ruc_2: str = ""
    historial_razon_social_2: str = ""
    historial_modalidad_2: str = ""
    historial_procedimiento_2: str = ""
    historial_fecha_emision_2: str = ""
    historial_fecha_venc_2: str = ""
    historial_fecha_baja_2: str = ""


def ensure_data_dirs(base_dir: Path) -> Path:
    entrada_dir = base_dir / "data" / "entrada_data"
    entrada_dir.mkdir(parents=True, exist_ok=True)
    return entrada_dir


def resolve_input_excel(entrada_dir: Path, explicit_path: str = "") -> Path:
    if explicit_path:
        path = Path(explicit_path).expanduser()
        if not path.is_absolute():
            path = entrada_dir.parent.parent / path
        if not path.exists():
            raise FileNotFoundError(f"No se encontro el archivo Excel configurado: {path}")
        return path

    candidates = sorted(
        [
            path
            for path in entrada_dir.glob("*.xlsx")
            if path.is_file() and not path.name.startswith("~$")
        ],
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError(
            f"No se encontro ningun .xlsx en {entrada_dir}. "
            "Coloca el archivo de entrada con la columna NRO DOCUMENTO."
        )
    return candidates[0]


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().upper()
    text = re.sub(r"\s+", " ", text)
    return text


def _cell_to_text_preserving_zeros(cell) -> str:
    value = cell.value
    if value is None:
        return ""

    if isinstance(value, str):
        return value.strip()

    number_format = str(cell.number_format or "")
    if isinstance(value, int):
        digits = str(value)
    elif isinstance(value, float) and value.is_integer():
        digits = str(int(value))
    else:
        return str(value).strip()

    # Si Excel guardo el nro de documento como numero con formato 00000000, recuperamos ceros de la mascara.
    if re.fullmatch(r"0+", number_format):
        return digits.zfill(len(number_format))
    return digits


def load_input_records(input_excel: Path, logger: logging.Logger) -> list[InputRecord]:
    workbook = load_workbook(input_excel, read_only=False, data_only=True)
    try:
        sheet = workbook.active

        header_cells = next(sheet.iter_rows(min_row=1, max_row=1), None)
        if not header_cells:
            raise ValueError("El Excel de entrada no tiene cabecera")

        headers = {_normalize_header(cell.value): index for index, cell in enumerate(header_cells)}
        document_index = headers.get("NRO DOCUMENTO")
        if document_index is None:
            document_index = headers.get("DNI")
        name_index = headers.get("APELLIDOS Y NOMBRES")

        if document_index is None:
            raise ValueError(
                "El Excel debe contener al menos la columna NRO DOCUMENTO "
                "(se acepta DNI como compatibilidad)."
            )

        if name_index is None:
            logger.warning(
                "La columna 'APELLIDOS Y NOMBRES' no existe en el Excel de entrada; se continuara usando solo NRO DOCUMENTO."
            )

        records: list[InputRecord] = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            nro_documento = _cell_to_text_preserving_zeros(row[document_index])
            apellidos_nombres = _cell_to_text_preserving_zeros(row[name_index]) if name_index is not None else ""
            if not nro_documento and not apellidos_nombres:
                continue
            if not nro_documento:
                logger.warning("[FILA %s] Registro omitido: NRO DOCUMENTO vacio", row_number)
                continue
            records.append(
                InputRecord(
                    row_number=row_number,
                    nro_documento=nro_documento,
                    apellidos_nombres=apellidos_nombres,
                )
            )

        logger.info("Excel de entrada cargado: %s | registros=%s", input_excel, len(records))
        return records
    finally:
        workbook.close()


def write_search_results(
    output_dir: Path,
    results: list[SearchResult],
    logger: logging.Logger,
    filename_prefix: str = "RB_GADSOCarnetSUCAMEC",
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{filename_prefix}_{datetime.now():%d.%m.%y_%H.%M.%S}.xlsx"

    workbook = Workbook()
    try:
        sheet = workbook.active
        sheet.title = "resultados"
        sheet.append(OUTPUT_FIELDS)

        for result in results:
            sheet.append([getattr(result, field) for field in OUTPUT_FIELDS])

        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            column_letter = column_cells[0].column_letter
            sheet.column_dimensions[column_letter].width = min(max(12, max_length + 2), 35)
            for cell in column_cells:
                cell.number_format = "@"

        workbook.save(output_path)
        logger.info("Resultado %s guardado en %s", filename_prefix, output_path)
        return output_path
    finally:
        workbook.close()
